import csv
import openpyxl

def check_values(input_csv, excel_file):
    found = []
    not_found = []
    values = []
    with open(input_csv) as f:
        reader = csv.reader(f)
        next(reader) # skip header
        values = [row[0] for row in reader]
    workbook = openpyxl.load_workbook(excel_file)
    for sheet in workbook:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value in values:
                    found.append((value, sheet.title))
                    values.remove(value)
    not_found = values
    return found, not_found

if __name__ == "__main__":
    input_csv = "input.csv"
    excel_file = "data.xlsx"
    output_file = "output.xlsx"
    found, not_found = check_values(input_csv, excel_file)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Results"
    worksheet.append(["Found/Not Found", "Value", "Sheet"])
    green_fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    for value, sheet_name in found:
        worksheet.append(["Found", value, sheet_name])
        cell = worksheet[f"A{worksheet.max_row}"]
        cell.fill = green_fill
    red_fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for value in not_found:
        worksheet.append(["Not Found", value, sheet_name])
        cell = worksheet[f"A{worksheet.max_row}"]
        cell.fill = red_fill
    workbook.save(output_file)
